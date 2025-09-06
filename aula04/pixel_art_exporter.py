import argparse
import csv
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def ensure_hex(rgb):
    r,g,b = rgb
    return f"#{r:02X}{g:02X}{b:02X}"

def image_to_matrix(path, size=(32,32), palette_colors=64, background=(255,255,255)):
    # Abre imagem com alpha e compõe sobre background
    img = Image.open(path).convert("RGBA")
    bg = Image.new("RGBA", img.size, background + (255,))
    bg.paste(img, mask=img.split()[3])  # usa alpha como máscara
    img_rgb = bg.convert("RGB")

    # Redimensiona com NEAREST para pixel-art
    img_small = img_rgb.resize(size, Image.NEAREST)

    # Quantiza a paleta com mais cores
    img_pal = img_small.convert("P", palette=Image.ADAPTIVE, colors=palette_colors)
    img_final = img_pal.convert("RGB")

    w,h = img_final.size
    pixels = list(img_final.getdata())

    matrix = []
    for r in range(h):
        row = []
        for c in range(w):
            row.append(ensure_hex(pixels[r*w + c]))
        matrix.append(row)
    return matrix

def write_csv(matrix, out_csv):
    with open(out_csv, "w", newline="") as f:
        writer = csv.writer(f)
        for row in matrix:
            writer.writerow(row)

def write_xlsx(matrix, out_xlsx, set_cell_fill=True):
    wb = Workbook()
    ws = wb.active
    rows = len(matrix)
    cols = len(matrix[0]) if rows>0 else 0

    for r_idx, row in enumerate(matrix, start=1):
        for c_idx, hexcol in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=hexcol)
            if set_cell_fill:
                color_argb = "FF" + hexcol.lstrip("#").upper()
                cell.fill = PatternFill(fill_type="solid", start_color=color_argb, end_color=color_argb)

    for c in range(1, cols+1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = 2.5

    for r in range(1, rows+1):
        ws.row_dimensions[r].height = 15

    wb.save(out_xlsx)

def make_preview_png(matrix, out_png, scale=16):
    h = len(matrix)
    w = len(matrix[0]) if h>0 else 0
    img = Image.new("RGB", (w, h))
    for r in range(h):
        for c in range(w):
            hexcol = matrix[r][c]
            rgb = tuple(int(hexcol[i:i+2], 16) for i in (1,3,5))
            img.putpixel((c, r), rgb)
    img = img.resize((w*scale, h*scale), Image.NEAREST)
    img.save(out_png)

def main():
    p = argparse.ArgumentParser(description="Converte imagem em pixel-art mantendo o tamanho, mas com mais cores.")
    p.add_argument("input", help="Arquivo de entrada (imagem).")
    p.add_argument("--size", type=int, nargs=2, metavar=("W","H"), default=(32,32),
                   help="Tamanho da pixel-art em pixels. Ex: --size 32 32")
    p.add_argument("--colors", type=int, default=64, help="Número de cores na paleta (ex: 16,32,64,128).")
    p.add_argument("--csv", default="pixelart.csv", help="Arquivo CSV de saída.")
    p.add_argument("--xlsx", default="PixelArt-Desenho.xlsx", help="Arquivo XLSX de saída.")
    p.add_argument("--preview", default="preview.png", help="PNG ampliado para visualizar.")
    p.add_argument("--no-preview", action="store_true", help="Não gerar preview PNG.")
    args = p.parse_args()

    size = tuple(args.size)
    matrix = image_to_matrix(args.input, size=size, palette_colors=args.colors)

    write_csv(matrix, args.csv)
    write_xlsx(matrix, args.xlsx)
    if not args.no_preview:
        make_preview_png(matrix, args.preview, scale=16)

    print(f"Gerado CSV: {args.csv}")
    print(f"Gerado XLSX: {args.xlsx}")
    if not args.no_preview:
        print(f"Gerado preview PNG: {args.preview}")

if __name__ == "__main__":
    main()
